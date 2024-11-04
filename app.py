
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
from io import BytesIO

st.title("Budget Data Extractor")

# Upload the template file
template_file = st.file_uploader("Upload Template File", type="xlsx")
uploaded_files = st.file_uploader("Upload Data Files", type="xlsx", accept_multiple_files=True)

if template_file and uploaded_files:
    # Load the template file and select the "Sheet1" sheet
    template_wb = load_workbook(template_file, data_only=True)
    template_sheet = template_wb['Sheet1']

    for data_file in uploaded_files:
        # Load the uploaded data workbook
        data_wb = load_workbook(data_file, data_only=True)
        data_sheet = data_wb['Page 1']

        # Initialize a dictionary for extracted data
        extracted_data = {}

        # Loop over each row in the F column to build the map and extract values
        for row in range(2, template_sheet.max_row + 1):
            cell_reference = template_sheet[f'F{row}'].value  # Get cell ref from column F
            
            if cell_reference:
                target_field = f'Field_{row - 1}'
                # Extract data from the specified cell in data workbook
                extracted_data[target_field] = data_sheet[cell_reference].value

                # Write extracted data into the corresponding cell in the template
                target_cell = f'G{row}'  # Adjust if needed for another location in template
                template_sheet[target_cell] = extracted_data[target_field]

        # Save updated template with a unique name for each upload
        data_file_name = os.path.splitext(data_file.name)[0]
        output_name = f"Formatted {data_file_name}.xlsx"
        output = BytesIO()
        template_wb.save(output)
        output.seek(0)

        st.download_button(
            label=f"Download {output_name}",
            data=output,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
